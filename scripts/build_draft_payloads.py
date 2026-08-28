#!/usr/bin/env python3
"""Build per-recipient Outlook draft payloads from a CSV or XLSX file.

This helper only creates JSON payloads. It never sends email and does not connect
to Outlook or Graph.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
from collections import defaultdict
from pathlib import Path
from string import Template


DEFAULT_SUBJECT = "[Action Required] Outstanding items - {{Count}} pending"
DEFAULT_TEMPLATE = Path(__file__).parent / "templates" / "default.html"
STATUS_COLUMNS = ("status", "state", "outcome", "decision")
OPEN_VALUES = ("open", "outstanding", "incomplete", "pending", "action required", "not started", "blocked")


def normalize_key(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def find_column(headers: list[str], preferred: str | None, candidates: tuple[str, ...]) -> str | None:
    if preferred:
        match = next((h for h in headers if h.lower() == preferred.lower()), None)
        if match:
            return match
        raise SystemExit(f"Column not found: {preferred}. Available: {', '.join(headers)}")

    normalized = {normalize_key(h): h for h in headers}
    for candidate in candidates:
        found = normalized.get(normalize_key(candidate))
        if found:
            return found
    return None


def render_text(template: str, values: dict[str, str]) -> str:
    rendered = template
    for key, value in values.items():
        rendered = rendered.replace("{{" + key + "}}", value)
    return rendered


def first_name(display: str, recipient: str) -> str:
    value = (display or recipient).strip()
    if "@" in value:
        value = value.split("@", 1)[0]
    value = value.replace(".", " ").replace("_", " ").strip()
    return value.split()[0].capitalize() if value else "there"


def build_table(rows: list[dict[str, str]], columns: list[str]) -> str:
    header = "".join(f"<th>{html.escape(col)}</th>" for col in columns)
    body_rows = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(row.get(col, '') or '')}</td>" for col in columns)
        body_rows.append(f"<tr>{cells}</tr>")
    return (
        "<table border=\"1\" cellpadding=\"6\" cellspacing=\"0\" "
        "style=\"border-collapse:collapse;border-color:#d0d7de\">"
        f"<thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"
    )


def should_keep(row: dict[str, str], status_column: str | None, include_all: bool) -> bool:
    if include_all or not status_column:
        return True
    status = (row.get(status_column) or "").strip().lower()
    return not status or any(value in status for value in OPEN_VALUES)


def read_rows(input_path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    if input_path.suffix.lower() == ".csv":
        return list(csv.DictReader(input_path.open(newline="", encoding="utf-8-sig")))

    if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("XLSX input requires openpyxl. Install it or export the workbook to CSV.") from exc

        workbook = load_workbook(input_path, data_only=True, read_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not raw_headers:
            return []
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        rows: list[dict[str, str]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {
                header: "" if value is None else str(value)
                for header, value in zip(headers, values)
                if header
            }
            if any(value for value in row.values()):
                rows.append(row)
        return rows

    raise SystemExit(f"Unsupported input type: {input_path.suffix}. Use .csv or .xlsx.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build per-recipient email draft payloads from CSV/XLSX.")
    parser.add_argument("--input", required=True, help="Input CSV path.")
    parser.add_argument("--output", required=True, help="Output JSON path.")
    parser.add_argument("--recipient-column", help="Recipient/alias/email column.")
    parser.add_argument("--display-name-column", help="Optional display-name column.")
    parser.add_argument("--status-column", help="Optional status column used to filter open rows.")
    parser.add_argument("--item-columns", nargs="*", help="Columns to include in the action-item table. Defaults to all non-recipient columns.")
    parser.add_argument("--subject", default=DEFAULT_SUBJECT, help="Subject template. Supports {{Count}} and {{Deadline}}.")
    parser.add_argument("--deadline", default="", help="Optional deadline text/date.")
    parser.add_argument("--sender-name", default="Parker", help="Sender display name used in template.")
    parser.add_argument("--sheet", help="Optional worksheet name for XLSX input.")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="HTML template path.")
    parser.add_argument("--include-all", action="store_true", help="Do not filter by status.")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    rows = read_rows(input_path, args.sheet)
    if not rows:
        raise SystemExit("No rows found in input CSV.")

    headers = list(rows[0].keys())
    recipient_column = find_column(headers, args.recipient_column, ("Owner", "Alias", "Email", "Reviewer", "PoC", "AssignedTo", "Assignee"))
    if not recipient_column:
        raise SystemExit(f"Could not infer recipient column. Available: {', '.join(headers)}")

    display_column = find_column(headers, args.display_name_column, ("DisplayName", "Name", "OwnerName", "ReviewerName"))
    status_column = find_column(headers, args.status_column, STATUS_COLUMNS)
    item_columns = args.item_columns or [h for h in headers if h not in {recipient_column, display_column}]

    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    display_names: dict[str, str] = {}
    skipped = 0
    for row in rows:
        if not should_keep(row, status_column, args.include_all):
            skipped += 1
            continue
        recipient = (row.get(recipient_column) or "").strip().rstrip("*")
        if not recipient:
            skipped += 1
            continue
        grouped[recipient].append(row)
        if display_column and row.get(display_column):
            display_names[recipient] = row[display_column]

    template = Path(args.template).expanduser().read_text(encoding="utf-8")
    payloads = []
    for recipient, recipient_rows in sorted(grouped.items(), key=lambda item: item[0].lower()):
        count = len(recipient_rows)
        values = {
            "FirstName": html.escape(first_name(display_names.get(recipient, ""), recipient)),
            "Count": str(count),
            "PluralSuffix": "" if count == 1 else "s",
            "Deadline": html.escape(args.deadline),
            "DeadlinePhrase": f" by <strong>{html.escape(args.deadline)}</strong>" if args.deadline else "",
            "SenderName": html.escape(args.sender_name),
            "ItemsTable_Html": build_table(recipient_rows, item_columns),
        }
        subject = render_text(args.subject, {"Count": str(count), "Deadline": args.deadline})
        payloads.append({
            "to": [recipient],
            "subject": subject,
            "contentType": "HTML",
            "body": render_text(template, values),
            "rowCount": count,
        })

    output = {
        "draftCount": len(payloads),
        "rowCount": sum(len(v) for v in grouped.values()),
        "skippedRowCount": skipped,
        "recipientColumn": recipient_column,
        "statusColumn": status_column,
        "itemColumns": item_columns,
        "drafts": payloads,
    }
    Path(args.output).expanduser().write_text(json.dumps(output, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
